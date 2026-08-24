"""Inventory export.

Two outputs, one source of truth (``schema.INVENTORY_FIELDS``):

- ``export_csv``  — plain CSV, the reliable path into Google Sheets.
- ``export_xlsx`` — a working workbook with data validation, conditional
  formatting, a dashboard, a field guide, and the comparable evidence.

openpyxl is imported lazily so the Telegram bot never needs it installed.
"""

from __future__ import annotations

import csv
from datetime import datetime
from pathlib import Path

from estate.repository import CompRepository, ItemRepository
from estate.schema import (
    CONDITIONS,
    FIELD_LABELS,
    FIELDS_BY_KEY,
    GROUPS,
    INVENTORY_FIELDS,
    STATUS_ORDER,
    ApprovalStatus,
    ReviewStatus,
    WebsiteStatus,
)

COMP_COLUMNS = [
    "Item ID", "Platform", "Title", "Sold or Active", "Price", "Price Type",
    "Shipping", "Total", "Condition", "Observed Date", "Location", "Relevance",
    "Similarities", "Differences", "Placeholder?", "Needs Confirmation?", "URL",
]


def _cell(item, key: str):
    value = getattr(item, key, None)
    if value is None:
        return ""
    if isinstance(value, bool):
        return "Yes" if value else "No"
    if isinstance(value, (list, tuple)):
        return "\n".join(str(v) for v in value)
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    return value


def rows(session) -> list:
    items = ItemRepository(session).all()
    return [[_cell(i, f.key) for f in INVENTORY_FIELDS] for i in items]


# ---------------------------------------------------------------------------
# CSV
# ---------------------------------------------------------------------------

def export_csv(session, out_path: Path | str) -> Path:
    out = Path(out_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    with out.open("w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow(FIELD_LABELS)
        for row in rows(session):
            w.writerow(row)
    return out


def export_comps_csv(session, out_path: Path | str) -> Path:
    out = Path(out_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    comps_repo = CompRepository(session)
    with out.open("w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow(COMP_COLUMNS)
        for item in ItemRepository(session).all():
            for c in comps_repo.for_item(item.item_id):
                total = float(c.price or 0) + float(c.shipping_amount or 0)
                w.writerow([
                    item.item_id, c.platform, c.title,
                    "Sold" if c.is_sold else "Active",
                    c.price, c.price_type or "exact",
                    c.shipping_amount, round(total, 2), c.condition,
                    c.observed_date, c.location, c.relevance,
                    c.similarities, c.differences,
                    "YES" if c.is_placeholder else "",
                    "YES" if c.needs_confirmation else "", c.url,
                ])
    return out


# ---------------------------------------------------------------------------
# XLSX
# ---------------------------------------------------------------------------

def export_xlsx(session, out_path: Path | str, max_rows: int = 500) -> Path:
    from openpyxl import Workbook
    from openpyxl.formatting.rule import CellIsRule, FormulaRule
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    out = Path(out_path)
    out.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    body_font = Font(name="Arial", size=10)
    group_fills = {
        "Identification": "2F4858", "Physical": "33658A", "Logistics": "55828B",
        "Research": "86BBD8", "Pricing": "F6AE2D", "Marketing": "9A7AA0",
        "Workflow": "6B705C", "Outcome": "3D5A80",
    }
    input_fill = PatternFill("solid", fgColor="FFFDE7")

    # -- Inventory ----------------------------------------------------------
    ws = wb.active
    ws.title = "Inventory"

    for col, f in enumerate(INVENTORY_FIELDS, start=1):
        c = ws.cell(row=1, column=col, value=f.group)
        c.font = Font(name="Arial", bold=True, size=8, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=group_fills.get(f.group, "555555"))
        c.alignment = Alignment(horizontal="center")

        h = ws.cell(row=2, column=col, value=f.label)
        h.font = header_font
        h.fill = PatternFill("solid", fgColor=group_fills.get(f.group, "555555"))
        h.alignment = Alignment(vertical="center", wrap_text=True)
        if f.note:
            h.comment = None  # notes live in the Field Guide sheet

        width = {"longtext": 42, "urls": 34, "money": 14, "date": 14,
                 "number": 12, "bool": 12, "percent": 12}.get(f.kind, 20)
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.freeze_panes = "C3"
    ws.auto_filter.ref = f"A2:{get_column_letter(len(INVENTORY_FIELDS))}2"

    data = rows(session)
    for r, row in enumerate(data, start=3):
        for col, value in enumerate(row, start=1):
            c = ws.cell(row=r, column=col, value=value)
            c.font = body_font
            kind = INVENTORY_FIELDS[col - 1].kind
            if kind == "money":
                c.number_format = '$#,##0;($#,##0);-'
            elif kind == "percent":
                c.number_format = "0.0%"
            elif kind in ("longtext", "urls"):
                c.alignment = Alignment(wrap_text=True, vertical="top")

    last = max(3, len(data) + 2)
    end = max(last, max_rows)

    # -- Validation ---------------------------------------------------------
    def add_choice_validation(key: str, choices, prompt: str) -> None:
        idx = [f.key for f in INVENTORY_FIELDS].index(key) + 1
        letter = get_column_letter(idx)
        dv = DataValidation(
            type="list", formula1='"{}"'.format(",".join(choices)), allow_blank=True,
            showDropDown=False, errorTitle="Not an allowed value",
            error=prompt, promptTitle=FIELDS_BY_KEY[key].label, prompt=prompt,
        )
        ws.add_data_validation(dv)
        dv.add("%s3:%s%d" % (letter, letter, end))

    add_choice_validation("status", STATUS_ORDER, "Pick the current lifecycle status.")
    add_choice_validation("condition", CONDITIONS, "Pick the closest condition grade.")
    add_choice_validation("review_status", [s.value for s in ReviewStatus],
                          "Where this item sits in review.")
    add_choice_validation("approval_status", [s.value for s in ApprovalStatus],
                          "A human must set this before anything is published.")
    add_choice_validation("website_status", [s.value for s in WebsiteStatus],
                          "Controls whether the item appears on the catalogue site.")

    for key in ("ownership_approval", "shipping_feasible", "pickup_required"):
        add_choice_validation(key, ["Yes", "No"], "Yes or No.")

    # -- Conditional formatting --------------------------------------------
    keys = [f.key for f in INVENTORY_FIELDS]
    status_col = get_column_letter(keys.index("status") + 1)
    conf_col = get_column_letter(keys.index("pricing_confidence") + 1)
    approval_col = get_column_letter(keys.index("approval_status") + 1)
    current_col = get_column_letter(keys.index("current_price") + 1)
    floor_col = get_column_letter(keys.index("floor_price") + 1)
    full = "A3:%s%d" % (get_column_letter(len(INVENTORY_FIELDS)), end)

    ws.conditional_formatting.add(
        full,
        FormulaRule(formula=[f'${status_col}3="Sold"'],
                    fill=PatternFill("solid", fgColor="D8F3DC"), stopIfTrue=False),
    )
    ws.conditional_formatting.add(
        full,
        FormulaRule(formula=[f'${status_col}3="Needs Review"'],
                    fill=PatternFill("solid", fgColor="FFF3CD"), stopIfTrue=False),
    )
    ws.conditional_formatting.add(
        full,
        FormulaRule(formula=[f'${status_col}3="Removed"'],
                    fill=PatternFill("solid", fgColor="EEEEEE"), stopIfTrue=False),
    )
    # Loud: a price at or below its floor, and unapproved items with a price.
    ws.conditional_formatting.add(
        "%s3:%s%d" % (current_col, current_col, end),
        FormulaRule(
            formula=[f'AND(${current_col}3<>"",${floor_col}3<>"",${current_col}3<=${floor_col}3)'],
            fill=PatternFill("solid", fgColor="F8D7DA"),
            font=Font(name="Arial", size=10, bold=True, color="9C0006"),
        ),
    )
    ws.conditional_formatting.add(
        "%s3:%s%d" % (conf_col, conf_col, end),
        CellIsRule(operator="equal", formula=['"Insufficient Evidence"'],
                   fill=PatternFill("solid", fgColor="F8D7DA")),
    )
    ws.conditional_formatting.add(
        "%s3:%s%d" % (conf_col, conf_col, end),
        CellIsRule(operator="equal", formula=['"Low"'],
                   fill=PatternFill("solid", fgColor="FFE5B4")),
    )
    ws.conditional_formatting.add(
        "%s3:%s%d" % (approval_col, approval_col, end),
        CellIsRule(operator="equal", formula=['"Approved"'],
                   fill=PatternFill("solid", fgColor="D8F3DC")),
    )

    # Highlight the columns a human is expected to fill in.
    for key in ("ownership_approval", "approval_status", "review_status",
                "approved_pickup_price", "buyer", "final_sale_price", "notes"):
        idx = keys.index(key) + 1
        for r in range(3, end + 1):
            ws.cell(row=r, column=idx).fill = input_fill

    # -- Dashboard ----------------------------------------------------------
    dash = wb.create_sheet("Dashboard", 0)
    dash.column_dimensions["A"].width = 34
    dash.column_dimensions["B"].width = 18
    dash["A1"] = "Estate Sale — Dashboard"
    dash["A1"].font = Font(name="Arial", bold=True, size=16)
    dash["A2"] = "Yellow cells on the Inventory tab are for a human to fill in."
    dash["A2"].font = Font(name="Arial", size=9, italic=True)

    inv_status = "Inventory!$%s$3:$%s$%d" % (status_col, status_col, end)
    price_range = "Inventory!$%s$3:$%s$%d" % (current_col, current_col, end)
    final_col = get_column_letter(keys.index("final_sale_price") + 1)
    final_range = "Inventory!$%s$3:$%s$%d" % (final_col, final_col, end)
    proceeds_col = get_column_letter(keys.index("actual_proceeds") + 1)
    proceeds_range = "Inventory!$%s$3:$%s$%d" % (proceeds_col, proceeds_col, end)
    id_col = get_column_letter(keys.index("item_id") + 1)
    id_range = "Inventory!$%s$3:$%s$%d" % (id_col, id_col, end)

    dash["A4"] = "Total items"
    dash["B4"] = f"=COUNTA({id_range})"
    row = 5
    for status in STATUS_ORDER:
        dash.cell(row=row, column=1, value=status)
        dash.cell(row=row, column=2, value=f'=COUNTIF({inv_status},"{status}")')
        row += 1

    row += 1
    dash.cell(row=row, column=1, value="Listed value (current prices)").font = Font(
        name="Arial", bold=True)
    dash.cell(row=row, column=2,
              value=f'=SUMIF({inv_status},"Listed",{price_range})'
              ).number_format = '$#,##0'
    row += 1
    dash.cell(row=row, column=1, value="Gross sold")
    dash.cell(row=row, column=2, value=f"=SUM({final_range})").number_format = '$#,##0'
    row += 1
    dash.cell(row=row, column=1, value="Net proceeds")
    dash.cell(row=row, column=2, value=f"=SUM({proceeds_range})").number_format = '$#,##0'
    row += 1
    dash.cell(row=row, column=1, value="Average sale price")
    dash.cell(row=row, column=2,
              value=f"=IFERROR(AVERAGEIF({final_range},\">0\"),0)").number_format = '$#,##0'
    row += 2
    dash.cell(row=row, column=1, value="Awaiting human approval").font = Font(
        name="Arial", bold=True, color="9C0006")
    dash.cell(row=row, column=2,
              value='=COUNTIF(Inventory!$%s$3:$%s$%d,"Pending")'
                    % (approval_col, approval_col, end))
    row += 1
    dash.cell(row=row, column=1, value="Low / insufficient pricing confidence")
    dash.cell(row=row, column=2,
              value='=COUNTIF(Inventory!$%s$3:$%s$%d,"Low")'
                    '+COUNTIF(Inventory!$%s$3:$%s$%d,"Insufficient Evidence")'
                    % (conf_col, conf_col, end, conf_col, conf_col, end))

    # -- Field guide --------------------------------------------------------
    guide = wb.create_sheet("Field Guide")
    guide.column_dimensions["A"].width = 18
    guide.column_dimensions["B"].width = 32
    guide.column_dimensions["C"].width = 14
    guide.column_dimensions["D"].width = 70
    for col, title in enumerate(["Section", "Column", "Type", "Notes / allowed values"], 1):
        c = guide.cell(row=1, column=col, value=title)
        c.font = header_font
        c.fill = PatternFill("solid", fgColor="2F4858")
    r = 2
    for group in GROUPS:
        for f in INVENTORY_FIELDS:
            if f.group != group:
                continue
            guide.cell(row=r, column=1, value=f.group)
            guide.cell(row=r, column=2, value=f.label)
            guide.cell(row=r, column=3, value=f.kind)
            note = f.note
            if f.choices:
                note = (note + " " if note else "") + "One of: " + ", ".join(f.choices)
            cell = guide.cell(row=r, column=4, value=note)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            r += 1

    # -- Comparable evidence ------------------------------------------------
    comps_ws = wb.create_sheet("Comparable Evidence")
    for col, title in enumerate(COMP_COLUMNS, 1):
        c = comps_ws.cell(row=1, column=col, value=title)
        c.font = header_font
        c.fill = PatternFill("solid", fgColor="86BBD8")
        comps_ws.column_dimensions[get_column_letter(col)].width = (
            46 if title in ("Title", "URL", "Similarities", "Differences") else 15
        )
    comps_repo = CompRepository(session)
    r = 2
    for item in ItemRepository(session).all():
        for c in comps_repo.for_item(item.item_id):
            # Column order: Item ID(1) Platform(2) Title(3) Sold or Active(4)
            # Price(5) Price Type(6) Shipping(7) Total(8) Condition(9)
            # Observed Date(10) Location(11) Relevance(12) Similarities(13)
            # Differences(14) Placeholder?(15) Needs Confirmation?(16) URL(17)
            values = [
                item.item_id, c.platform, c.title, "Sold" if c.is_sold else "Active",
                c.price, c.price_type or "exact", c.shipping_amount, None,
                c.condition, c.observed_date,
                c.location, c.relevance, c.similarities, c.differences,
                "YES" if c.is_placeholder else "",
                "YES" if c.needs_confirmation else "", c.url,
            ]
            for col, v in enumerate(values, 1):
                comps_ws.cell(row=r, column=col, value=v)
            comps_ws.cell(row=r, column=8, value="=E%d+G%d" % (r, r)).number_format = '$#,##0.00'
            comps_ws.cell(row=r, column=5).number_format = '$#,##0.00'
            comps_ws.cell(row=r, column=7).number_format = '$#,##0.00'
            if c.is_placeholder:
                for col in range(1, len(COMP_COLUMNS) + 1):
                    comps_ws.cell(row=r, column=col).fill = PatternFill(
                        "solid", fgColor="F8D7DA")
            elif c.needs_confirmation:
                for col in range(1, len(COMP_COLUMNS) + 1):
                    comps_ws.cell(row=r, column=col).fill = PatternFill(
                        "solid", fgColor="FFE5B4")
            r += 1
    comps_ws.freeze_panes = "A2"
    if r > 2:
        comps_ws.auto_filter.ref = "A1:%s%d" % (get_column_letter(len(COMP_COLUMNS)), r - 1)

    note = comps_ws.cell(row=r + 1, column=1,
                         value="Rows shaded red are PLACEHOLDER data and are not real "
                               "market evidence. Never price an item from them.")
    note.font = Font(name="Arial", bold=True, color="9C0006")

    wb.save(out)
    return out
